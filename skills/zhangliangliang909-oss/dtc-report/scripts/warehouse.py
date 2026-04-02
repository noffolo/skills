#!/usr/bin/env python3
"""
海外仓分析模块

包含：
- 仓库分布结构
- 入库柜量和出库件数分布（按国家）
- 美国入库柜量分布（自营/第三方）
- 美国出库件数分布（自营/第三方）
- Top10 亏损客户
"""

import openpyxl
from collections import defaultdict
import os
import glob

# 数据目录
BUSINESS_DATA_DIR = "/home/admin/.openclaw/workspace/agents/跨境电商财务分析-agent/data/1.业务和订单数据"

# 自营仓代码
SELF_RUN_WAREHOUSES = ['USCAEA02', 'USNJHM01']

# 仓库代码与地区映射
WAREHOUSE_REGION = {
    'USNJHM01': '美东',  # 新泽西美东自营仓
    'USCAEA02': '美西',  # 加州美西自营仓
    # 第三方仓库地区映射（根据仓库位置判断）
}

# 地区分类
US_REGIONS = ['美东', '美西', '美南', '美国中西部']

# 币种与国家映射
CURRENCY_COUNTRY = {
    'USD': '美国',
    'EUR': '欧洲',
    'GBP': '英国',
    'CAD': '加拿大',
    'AUD': '澳洲',
    '其他': '其他'
}


def read_warehouse_data(period='2026-Q1'):
    """读取海外仓数据
    
    返回：
    - 按国家分布的入库柜量和出库件数
    - 美国仓库分布（自营/第三方）
    - 各仓库入库柜量趋势
    - 各仓库出库件数趋势
    """
    current_month, cumulative_months = parse_period(period)
    
    result = {
        'by_country': defaultdict(lambda: {'inbound_containers': 0, 'outbound_pieces': 0}),
        'us_warehouse': {
            'self_run': {'inbound_containers': 0, 'outbound_pieces': 0},
            'third_party': {'inbound_containers': 0, 'outbound_pieces': 0},
            'self_run_by_region': {region: {'inbound_containers': 0, 'outbound_pieces': 0} for region in US_REGIONS},
            'third_party_by_region': {region: {'inbound_containers': 0, 'outbound_pieces': 0} for region in US_REGIONS}
        },
        'self_run_trend': {
            'USNJHM01': {'inbound': [], 'outbound': []},  # 美东自营仓
            'USCAEA02': {'inbound': [], 'outbound': []}   # 美西自营仓
        },
        'warehouse_trend': defaultdict(lambda: {'inbound': [], 'outbound': []}),
        'months': []
    }
    
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if not files:
        return result
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    # 找表头
    headers = {}
    for col in range(1, 30):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    warehouse_col = headers.get('目标仓库编码', 3)
    currency_col = headers.get('客户结算币种', 5)
    location_col = headers.get('仓库位置', 6)
    inbound_vol_col = headers.get('入库箱量', 8)
    outbound_pieces_col = headers.get('出库总件数', 19)
    month_col = headers.get('业务月份', 4)
    
    # 按月汇总
    monthly_data = defaultdict(lambda: defaultdict(lambda: {'inbound': 0, 'outbound': 0}))
    self_run_monthly = defaultdict(lambda: {'USNJHM01': {'inbound': 0, 'outbound': 0}, 'USCAEA02': {'inbound': 0, 'outbound': 0}})
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        warehouse = ws.cell(row=row, column=warehouse_col).value
        currency = ws.cell(row=row, column=currency_col).value
        month = ws.cell(row=row, column=month_col).value
        inbound_vol = ws.cell(row=row, column=inbound_vol_col).value
        outbound_pieces = ws.cell(row=row, column=outbound_pieces_col).value
        location = ws.cell(row=row, column=location_col).value
        
        if not month:
            continue
        
        # 处理 datetime 类型
        from datetime import datetime
        if isinstance(month, datetime):
            month_str = f"{month.year}-{month.month}"  # 不补零
        else:
            month_str = str(month)
            if '年' in month_str and '月' in month_str:
                month_str = month_str.replace('年', '-').replace('月', '').strip()
            elif '-' in month_str:
                pass
            else:
                continue
            
            # 转为整数去掉前导零
            parts = month_str.split('-')
            if len(parts) == 2:
                year = parts[0].strip()
                month_num = int(parts[1].strip())
                month_str = f"{year}-{month_num}"
            else:
                continue
        
        # 解析月份
        parts = month_str.split('-')
        if len(parts) >= 2:
            year = parts[0].strip()
            month_num = int(parts[1].strip())
        else:
            continue
        
        inbound = float(inbound_vol) if inbound_vol else 0
        outbound = float(outbound_pieces) if outbound_pieces else 0
        
        # 换算为柜量（1 柜=2TEU）
        inbound_containers = inbound / 2
        
        # 确定国家
        country = CURRENCY_COUNTRY.get(currency, '其他') if currency else '其他'
        
        # 累计数据：只统计 2026 年截至当期累计数据（Q1 报告只统计 2026 Q1）
        if year == '2026' and month_num in cumulative_months:
            result['by_country'][country]['inbound_containers'] += inbound_containers
            result['by_country'][country]['outbound_pieces'] += outbound
        
        # 美国仓库区分自营/第三方，并按地区分类（只统计 2026 Q1）
        if country == '美国' and warehouse and year == '2026' and month_num in cumulative_months:
            # 确定地区（根据仓库编码）
            region = None
            if warehouse:
                wh_str = str(warehouse)
                # 美东：NJ, NY, PA 等
                if 'NJ' in wh_str or 'NY' in wh_str or 'PA' in wh_str:
                    region = '美东'
                # 美西：CA, WA, OR 等
                elif 'CA' in wh_str or 'WA' in wh_str or 'OR' in wh_str:
                    region = '美西'
                # 美南：TX, GA, FL 等
                elif 'TX' in wh_str or 'GA' in wh_str or 'FL' in wh_str:
                    region = '美南'
                # 美国中西部：IL, OH, MI, IN 等
                elif 'IL' in wh_str or 'OH' in wh_str or 'MI' in wh_str or 'IN' in wh_str or 'CHI' in wh_str:
                    region = '美国中西部'
                # 其他美国仓库
                elif 'US' in wh_str:
                    region = '美东'  # 默认归到美东
            
            # 如果仓库编码无法判断，使用位置信息
            if not region and location:
                location_str = str(location)
                if '东' in location_str or 'NJ' in location_str or 'NY' in location_str:
                    region = '美东'
                elif '西' in location_str or 'CA' in location_str or 'LA' in location_str:
                    region = '美西'
                elif '南' in location_str or 'TX' in location_str or 'GA' in location_str:
                    region = '美南'
                elif '中西' in location_str or 'IL' in location_str or 'OH' in location_str or 'CHI' in location_str:
                    region = '美国中西部'
                else:
                    region = '美东'  # 默认归到美东
            
            if warehouse in SELF_RUN_WAREHOUSES:
                result['us_warehouse']['self_run']['inbound_containers'] += inbound_containers
                result['us_warehouse']['self_run']['outbound_pieces'] += outbound
                # 自营仓按地区统计
                if region and region in US_REGIONS:
                    result['us_warehouse']['self_run_by_region'][region]['inbound_containers'] += inbound_containers
                    result['us_warehouse']['self_run_by_region'][region]['outbound_pieces'] += outbound
            else:
                result['us_warehouse']['third_party']['inbound_containers'] += inbound_containers
                result['us_warehouse']['third_party']['outbound_pieces'] += outbound
                # 第三方仓按地区统计
                if region and region in US_REGIONS:
                    result['us_warehouse']['third_party_by_region'][region]['inbound_containers'] += inbound_containers
                    result['us_warehouse']['third_party_by_region'][region]['outbound_pieces'] += outbound
        
        # 月度趋势数据
        monthly_data[month_str][warehouse]['inbound'] += inbound_containers
        monthly_data[month_str][warehouse]['outbound'] += outbound
        
        # 自营仓月度数据
        if warehouse in SELF_RUN_WAREHOUSES:
            self_run_monthly[month_str][warehouse]['inbound'] += inbound_containers
            self_run_monthly[month_str][warehouse]['outbound'] += outbound
    
    # 整理趋势数据，取最近 6 个月
    all_months = sorted(monthly_data.keys())
    result['months'] = all_months[-6:] if len(all_months) > 6 else all_months
    
    warehouses = set()
    for m in result['months']:
        for wh in monthly_data[m].keys():
            warehouses.add(wh)
    
    for wh in warehouses:
        for m in result['months']:
            result['warehouse_trend'][wh]['inbound'].append(monthly_data[m][wh]['inbound'])
            result['warehouse_trend'][wh]['outbound'].append(monthly_data[m][wh]['outbound'])
    
    # 整理自营仓趋势数据（近 12 个月）
    all_months = sorted(self_run_monthly.keys(), key=lambda m: (int(m.split('-')[0]), int(m.split('-')[1])))
    result['months'] = all_months[-12:] if len(all_months) > 12 else all_months
    
    for wh in SELF_RUN_WAREHOUSES:
        for m in result['months']:
            if m in self_run_monthly and wh in self_run_monthly[m]:
                result['self_run_trend'][wh]['inbound'].append(self_run_monthly[m][wh]['inbound'])
                result['self_run_trend'][wh]['outbound'].append(self_run_monthly[m][wh]['outbound'])
            else:
                result['self_run_trend'][wh]['inbound'].append(0)
                result['self_run_trend'][wh]['outbound'].append(0)
    
    return result


def read_loss_customers(period='2026-Q1', top_n=10):
    """读取 Top10 亏损客户
    
    返回：
    - [(客户名称，毛利，仓库), ...]
    """
    files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*客户仓库*.xlsx'))
    if not files:
        return []
    
    f = files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    headers = {}
    for col in range(1, 30):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    # 从所有业务收入明细读取客户毛利
    revenue_files = glob.glob(os.path.join(BUSINESS_DATA_DIR, '*所有业务收入明细*.xlsx'))
    if not revenue_files:
        return []
    
    f = revenue_files[0]
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    
    headers = {}
    for col in range(1, 20):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip()] = col
    
    customer_col = headers.get('委托客户名称', 4)
    month_col = headers.get('业务年月', 3)
    profit_col = headers.get('毛利', 6)
    segment_col = headers.get('业务分段分类_新.', 2)
    
    # 汇总客户毛利（只统计 D 段）
    customer_profit = defaultdict(float)
    
    current_month, cumulative_months = parse_period(period)
    
    for row in range(2, min(ws.max_row + 1, 50000)):
        customer = ws.cell(row=row, column=customer_col).value
        month = ws.cell(row=row, column=month_col).value
        profit = ws.cell(row=row, column=profit_col).value
        segment = ws.cell(row=row, column=segment_col).value
        
        if not customer or not month:
            continue
        
        # 只统计 D 段
        seg_str = str(segment).strip() if segment else ''
        if 'D' not in seg_str and '电商' not in seg_str and '集拼' not in seg_str:
            continue
        
        # 处理 datetime 类型
        from datetime import datetime
        if isinstance(month, datetime):
            month_str = f"{month.year}-{month.month}"  # 不补零
        else:
            month_str = str(month)
            if '年' in month_str and '月' in month_str:
                month_str = month_str.replace('年', '-').replace('月', '').strip()
            elif '-' in month_str:
                pass
            else:
                continue
            
            # 转为整数去掉前导零
            parts = month_str.split('-')
            if len(parts) == 2:
                year = parts[0].strip()
                month_num = int(parts[1].strip())
                month_str = f"{year}-{month_num}"
            else:
                continue
        
        if month_num not in cumulative_months:
            continue
        
        customer_profit[customer] += float(profit) if profit else 0
    
    # 筛选亏损客户（毛利<0）
    loss_customers = [(name, prof) for name, prof in customer_profit.items() if prof < 0]
    
    # 按毛利排序（从小到大，最亏损的在前）
    loss_customers.sort(key=lambda x: x[1])
    
    return loss_customers[:top_n]


def parse_period(period):
    """解析期间参数"""
    if period.upper().startswith('2026-Q'):
        quarter = int(period[-1])
        current_month = quarter * 3
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    if period.startswith('2026-'):
        month_part = period[5:]
        current_month = int(month_part)
        cumulative_months = list(range(1, current_month + 1))
        return current_month, cumulative_months
    
    return 3, [1, 2, 3]
